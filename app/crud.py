# app/crud.py
import os
import shutil
import uuid
import io
import pandas as pd
import openpyxl
from fastapi import UploadFile, HTTPException
from sqlalchemy.orm import Session, joinedload
from sqlalchemy.orm import Session, joinedload, selectinload
from sqlalchemy.sql import func
from typing import List, Dict, Any, Optional
from datetime import date
from . import models, schemas, security
try:
    import ifcopenshell
except ImportError:
    ifcopenshell = None
UPLOAD_DIRECTORY = os.path.join(os.getcwd(), "uploads")
def get_user(db: Session, user_id: int):
    return db.query(models.User).filter(models.User.id == user_id).first()
def get_user_by_email(db: Session, email: str):
    return db.query(models.User).filter(models.User.email == email).first()
def create_user(db: Session, user: schemas.UserCreate):
    hashed_password = security.get_password_hash(user.password)
    db_user = models.User(
        email=user.email, 
        hashed_password=hashed_password,
        company_id=user.company_id,
        role=user.role )
    db.add(db_user); db.commit(); db.refresh(db_user); return db_user
def get_company(db: Session, company_id: int):
    return db.query(models.Company).filter(models.Company.id == company_id).first()
def get_company_by_name(db: Session, name: str):
    return db.query(models.Company).filter(models.Company.name == name).first()
def create_company(db: Session, company: schemas.CompanyCreate):
    db_company = models.Company(name=company.name)
    db.add(db_company); db.commit(); db.refresh(db_company); return db_company

def create_concept(db: Session, concept: schemas.ConceptCreate):
    db_concept = models.Concept(**concept.model_dump())
    db.add(db_concept)
    db.commit()
    db.refresh(db_concept)
    return db_concept

def get_concepts(db: Session, company_id: int, skip: int = 0, limit: int = 100):
    return db.query(models.Concept).filter(models.Concept.company_id == company_id).offset(skip).limit(limit).all()

def get_concept(db: Session, concept_id: int):
    return db.query(models.Concept).filter(models.Concept.id == concept_id).first()

def update_concept(db: Session, concept_id: int, concept_update: schemas.ConceptUpdate):
    db_concept = get_concept(db, concept_id)
    if not db_concept: return None
    update_data = concept_update.model_dump(exclude_unset=True)
    for key, value in update_data.items(): setattr(db_concept, key, value)
    db.add(db_concept); db.commit(); db.refresh(db_concept); return db_concept

def import_concepts_from_contracts(db: Session, company_id: int):
    projects = db.query(models.Project).filter(models.Project.company_id == company_id).all()
    project_ids = [p.id for p in projects]
    
    if not project_ids:
        return 0        
    items = db.query(models.ContractItem).join(models.Contract).filter(
        models.Contract.project_id.in_(project_ids),
        models.ContractItem.is_group == False 
    ).all()    
    imported_count = 0    
    for item in items:
        code = item.clave or f"GEN-{uuid.uuid4().hex[:8]}"        
        existing = db.query(models.Concept).filter(
            models.Concept.company_id == company_id,
            models.Concept.code == code
        ).first()        
        if not existing:
            new_concept = models.Concept(
                code=code,
                description=item.concepto,
                unit=item.unidad or "un",
                unit_price=item.precio_unitario,
                company_id=company_id
            )
            db.add(new_concept)
            imported_count += 1
            
    db.commit()
    return imported_count

def delete_concept(db: Session, concept_id: int):
    db_concept = get_concept(db, concept_id)
    if not db_concept: return None
    db.delete(db_concept); db.commit(); return True
def authenticate_user(db: Session, email: str, password: str):
    user = get_user_by_email(db, email=email)
    if not user:
        return None
    if not security.verify_password(password, user.hashed_password):
        return None
    return user
def get_projects(db: Session, company_id: int, skip: int = 0, limit: int = 100):
    return db.query(models.Project).filter(models.Project.company_id == company_id).options(
        selectinload(models.Project.tasks),
        selectinload(models.Project.folders),
        selectinload(models.Project.work_items).selectinload(models.WorkItem.contracts),
        selectinload(models.Project.contracts).options(
            selectinload(models.Contract.contractor),
            selectinload(models.Contract.contract_items),
            selectinload(models.Contract.estimates) ),
        selectinload(models.Project.contractors)    ).offset(skip).limit(limit).all()
def get_project_details(db: Session, project_id: int, company_id: int):
    return db.query(models.Project).filter(
        models.Project.id == project_id,
        models.Project.company_id == company_id
    ).options(
        selectinload(models.Project.tasks),
        selectinload(models.Project.folders),
        selectinload(models.Project.work_items),
        selectinload(models.Project.contracts),
        selectinload(models.Project.estimates),
        selectinload(models.Project.contractors)
    ).first()
def delete_project_by_id(db: Session, project_id: int, company_id: int):
    db_project = db.query(models.Project).filter(
        models.Project.id == project_id, models.Project.company_id == company_id
    ).first()
    if not db_project: return None
    db.delete(db_project); db.commit(); return db_project
def update_project(db: Session, project_id: int, project_update: schemas.ProjectUpdate, company_id: int):
    db_project = db.query(models.Project).filter(
        models.Project.id == project_id, models.Project.company_id == company_id
    ).first()
    if not db_project: return None
    update_data = project_update.model_dump(exclude_none=True)
    for key, value in update_data.items(): setattr(db_project, key, value)
    db.add(db_project); db.commit(); db.refresh(db_project); return db_project
def create_project(db: Session, project: schemas.ProjectCreate):
    db_project = models.Project(**project.model_dump())
    db.add(db_project)
    db.commit()
    db.refresh(db_project)
    return db_project
def create_task(db: Session, task: schemas.TaskCreate, project_id: int, creator_id: int):
    db_task = models.Task(**task.model_dump(), project_id=project_id, creator_id=creator_id)
    db.add(db_task); db.commit(); db.refresh(db_task); return db_task
def get_task(db: Session, task_id: int):
    return db.query(models.Task).options(joinedload(models.Task.project)).filter(models.Task.id == task_id).first()
def update_task(db: Session, task_id: int, task_update: schemas.TaskUpdate):
    db_task = get_task(db, task_id=task_id)
    if not db_task: return None
    update_data = task_update.model_dump(exclude_none=True)
    for key, value in update_data.items(): setattr(db_task, key, value)
    db.add(db_task); db.commit(); db.refresh(db_task); return db_task
def delete_task(db: Session, task_id: int, user: models.User):
    db_task = get_task(db, task_id=task_id)
    if not db_task:
        raise HTTPException(status_code=404, detail="Tarea no encontrada")
    if user.role == 'admin' and db_task.project.company_id == user.company_id:
        db.delete(db_task)
        db.commit()
        return {"ok": True}
    if user.role == 'user' and db_task.creator_id == user.id:
        db.delete(db_task)
        db.commit()
        return {"ok": True}
    raise HTTPException(status_code=403, detail="No tiene permisos para eliminar esta tarea")

def recalculate_task_cost(db: Session, task_id: int):
    task = get_task(db, task_id)
    if not task: return
    
    total_cost = 0.0
    for tc in task.concepts:
        total_cost += tc.amount
    
    task.estimated_cost = total_cost
    db.add(task)
    db.commit()
    db.refresh(task)

def assign_concept_to_task(db: Session, task_id: int, concept_id: int, quantity: float):
    task = get_task(db, task_id)
    concept = get_concept(db, concept_id)
    if not task or not concept: return None    
    existing = db.query(models.TaskConcept).filter(models.TaskConcept.task_id == task_id, models.TaskConcept.concept_id == concept_id).first()
    if existing:
        existing.quantity = quantity
        existing.amount = quantity * concept.unit_price
        db.add(existing)
    else:
        new_tc = models.TaskConcept(
            task_id=task_id,
            concept_id=concept_id,
            quantity=quantity,
            amount=quantity * concept.unit_price
        )
        db.add(new_tc)
    
    db.commit()
    recalculate_task_cost(db, task_id)
    return task

def remove_concept_from_task(db: Session, task_id: int, concept_id: int):
    tc = db.query(models.TaskConcept).filter(models.TaskConcept.task_id == task_id, models.TaskConcept.concept_id == concept_id).first()
    if tc:
        db.delete(tc)
        db.commit()
        recalculate_task_cost(db, task_id)
        return True
    return False


def add_task_dependency(db: Session, predecessor_id: int, successor_id: int, type: str = "FS", lag: int = 0):
    if predecessor_id == successor_id:
        return False
    
    stmt = models.task_dependencies.select().where(
        (models.task_dependencies.c.predecessor_id == predecessor_id) & 
        (models.task_dependencies.c.successor_id == successor_id)
    )
    existing = db.execute(stmt).first()
    if existing:
        return True
        
    ins = models.task_dependencies.insert().values(
        predecessor_id=predecessor_id,
        successor_id=successor_id,
        type=type,
        lag=lag
    )
    db.execute(ins)
    db.commit()
    return True

def remove_task_dependency(db: Session, predecessor_id: int, successor_id: int):
    stmt = models.task_dependencies.delete().where(
        (models.task_dependencies.c.predecessor_id == predecessor_id) & 
        (models.task_dependencies.c.successor_id == successor_id)
    )
    db.execute(stmt)
    db.commit()
    return True

def get_root_tasks(db: Session, project_id: int):
    return db.query(models.Task).filter(
        models.Task.project_id == project_id, models.Task.parent_id == None
    ).all()

def get_project_tasks(db: Session, project_id: int):
    return db.query(models.Task).filter(models.Task.project_id == project_id).all()
def get_project_tasks_as_tree(db: Session, project_id: int) -> List[schemas.Task]:
    db_tasks = get_project_tasks(db=db, project_id=project_id)
    task_schema_map = {}
    for db_task in db_tasks:
        task_schema = schemas.Task.from_orm(db_task)
        task_schema.subtasks = []
        task_schema_map[task_schema.id] = task_schema    
    root_tasks = []
    for task_id, task in task_schema_map.items():
        if task.parent_id:
            parent_task = task_schema_map.get(task.parent_id)
            if parent_task:
                parent_task.subtasks.append(task)
        else:
            root_tasks.append(task)

    def _calculate_parent_dates(task: schemas.Task):
        if not task.subtasks:
            if not task.start_date or not task.end_date: return None, None
            return task.start_date, task.end_date
        min_start, max_end = date.max, date.min; has_valid_dates = False
        for subtask in task.subtasks:
            child_start, child_end = _calculate_parent_dates(subtask)
            if child_start and child_end:
                has_valid_dates = True
                if child_start < min_start: min_start = child_start
                if child_end > max_end: max_end = child_end
        if has_valid_dates:
            task.start_date, task.end_date = min_start, max_end
            return min_start, max_end
        else:
            return task.start_date, task.end_date
    for root_task in root_tasks: _calculate_parent_dates(root_task)
    return root_tasks
    
def import_tasks_from_excel(db: Session, project_id: int, file_contents: bytes):
    pass

def create_contractor(db: Session, contractor: schemas.ContractorCreate):
    db_contractor = models.Contractor(**contractor.model_dump())
    db.add(db_contractor); db.commit(); db.refresh(db_contractor); return db_contractor

def get_all_company_contractors(db: Session, company_id: int, skip: int = 0, limit: int = 100):
    # Esta función ahora busca contratistas a través de los proyectos de la compañía
    return db.query(models.Contractor)\
        .join(models.Project, models.Contractor.project_id == models.Project.id)\
        .filter(models.Project.company_id == company_id)\
        .offset(skip)\
        .limit(limit)\
        .all()

def get_contractors(db: Session, project_id: int, skip: int = 0, limit: int = 100):
    return db.query(models.Contractor).filter(models.Contractor.project_id == project_id).offset(skip).limit(limit).all()

def get_contractor(db: Session, contractor_id: int):
    return db.query(models.Contractor).filter(models.Contractor.id == contractor_id).first()

def get_contractor_by_razon_social(db: Session, razon_social: str, project_id: int):
    return db.query(models.Contractor).filter(
        models.Contractor.razon_social == razon_social, 
        models.Contractor.project_id == project_id
    ).first()

def update_contractor(db: Session, contractor_id: int, contractor_update: schemas.ContractorUpdate, user_company_id: int):
    db_contractor = db.query(models.Contractor).options(joinedload(models.Contractor.project)).filter(models.Contractor.id == contractor_id).first()
    if not db_contractor: return None
    if db_contractor.project.company_id != user_company_id:
        raise HTTPException(status_code=403, detail="No tiene permisos sobre este contratista")
    update_data = contractor_update.model_dump(exclude_none=True)
    for key, value in update_data.items(): setattr(db_contractor, key, value)
    db.add(db_contractor); db.commit(); db.refresh(db_contractor); return db_contractor

def delete_contractor(db: Session, contractor_id: int, user_company_id: int):
    db_contractor = db.query(models.Contractor).options(joinedload(models.Contractor.project)).filter(models.Contractor.id == contractor_id).first()
    if not db_contractor:
        raise HTTPException(status_code=404, detail="Contratista no encontrado")
    if db_contractor.project.company_id != user_company_id:
        raise HTTPException(status_code=403, detail="No tiene permisos sobre este contratista")

    # Verificar si el contratista tiene contratos asociados
    has_contracts = db.query(models.Contract).filter(models.Contract.contractor_id == contractor_id).first()
    if has_contracts:
        raise HTTPException(status_code=400, detail="No se puede eliminar el contratista porque tiene contratos asociados.")

    db.delete(db_contractor)
    db.commit()
    return {"ok": True}

def import_contractors_from_excel(db: Session, file_contents: bytes, project_id: int):
    workbook = openpyxl.load_workbook(io.BytesIO(file_contents))
    sheet = workbook.active    
    created_count = 0
    skipped_count = 0
    errors = []
    for i, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        try:
            razon_social = str(row[0]).strip() if row[0] else None
            if not razon_social:
                continue

            existing_contractor = get_contractor_by_razon_social(db, razon_social=razon_social, project_id=project_id)
            if existing_contractor:
                skipped_count += 1
                continue

            contractor_data = schemas.ContractorCreate(razon_social=razon_social, responsable=str(row[1]) if row[1] else None, telefono=str(row[2]) if row[2] else None, correo_electronico=str(row[3]) if row[3] else None, project_id=project_id)
            create_contractor(db=db, contractor=contractor_data); created_count += 1
        except Exception as e:
            errors.append(f"Fila {i}: Error inesperado - {e}")
            continue
    return {"message": f"{created_count} contratistas creados, {skipped_count} omitidos por duplicados.", "errors": errors}

def create_work_item(db: Session, work_item: schemas.WorkItemCreate, project_id: int):
    db_work_item = models.WorkItem(**work_item.model_dump(), project_id=project_id)
    db.add(db_work_item); db.commit(); db.refresh(db_work_item); return db_work_item
def get_project_work_items(db: Session, project_id: int, skip: int = 0, limit: int = 100):
    return db.query(models.WorkItem).filter(models.WorkItem.project_id == project_id).options(
        joinedload(models.WorkItem.contracts).options(
            selectinload(models.Contract.contract_items).options(
                selectinload(models.ContractItem.subitems).selectinload(models.ContractItem.subitems)
            ),
        )
    ).offset(skip).limit(limit).all()
def get_work_item(db: Session, work_item_id: int):
    return db.query(models.WorkItem).filter(models.WorkItem.id == work_item_id).first()
def update_work_item(db: Session, work_item_id: int, work_item_update: schemas.WorkItemUpdate):
    db_work_item = get_work_item(db, work_item_id=work_item_id)
    if not db_work_item: return None
    update_data = work_item_update.model_dump(exclude_none=True)
    for key, value in update_data.items(): setattr(db_work_item, key, value)
    db.add(db_work_item); db.commit(); db.refresh(db_work_item); return db_work_item
def delete_work_item(db: Session, work_item_id: int):
    db_work_item = get_work_item(db, work_item_id=work_item_id)
    if not db_work_item: return None
    db.delete(db_work_item); db.commit(); return db_work_item

def create_contract(db: Session, contract: schemas.ContractCreate, project_id: int):
    db_contract = models.Contract(**contract.model_dump(), project_id=project_id)
    db.add(db_contract); db.commit(); db.refresh(db_contract); return get_contract(db, db_contract.id)
def get_project_contracts(db: Session, project_id: int):
    return db.query(models.Contract).filter(models.Contract.project_id == project_id).options(
        joinedload(models.Contract.contractor),
        joinedload(models.Contract.work_item),
        joinedload(models.Contract.contract_items).options(
            joinedload(models.ContractItem.estimate_items)
        ),
        joinedload(models.Contract.estimates).options(
            joinedload(models.Estimate.estimate_items)
        )
    ).all()
def get_contract(db: Session, contract_id: int):
    return db.query(models.Contract).options(
        joinedload(models.Contract.project) 
    ).filter(models.Contract.id == contract_id).first()

def get_contract_by_number(db: Session, project_id: int, numero_contrato: str):
    return db.query(models.Contract).filter(models.Contract.project_id == project_id, models.Contract.numero_contrato == numero_contrato).first()

def get_contract_for_permission_check(db: Session, contract_id: int):

    return db.query(models.Contract).options(joinedload(models.Contract.project)).filter(models.Contract.id == contract_id).first()

def get_contract_details(db: Session, contract_id: int):
    return db.query(models.Contract).filter(models.Contract.id == contract_id).options(
        joinedload(models.Contract.contractor),
        joinedload(models.Contract.dms_folder),
        joinedload(models.Contract.work_item),       
        selectinload(models.Contract.contract_items).options(
            selectinload(models.ContractItem.subitems).selectinload(models.ContractItem.subitems),
            selectinload(models.ContractItem.estimate_items)
        ),
        joinedload(models.Contract.estimates).options(
            joinedload(models.Estimate.estimate_items).options(
                joinedload(models.EstimateItem.contract_item)
            )
        )
    ).first()
def update_contract(db: Session, contract_id: int, contract_update: schemas.ContractUpdate):  
    db_contract = db.query(models.Contract).filter(models.Contract.id == contract_id).first()
    if not db_contract: return None
    update_data = contract_update.model_dump(exclude_none=True)
    if 'work_item_id' in update_data:
        work_item_id = update_data.pop('work_item_id')
        db_contract.work_item_id = work_item_id
    for key, value in update_data.items():
        setattr(db_contract, key, value)
    db.add(db_contract); db.commit(); db.refresh(db_contract); return get_contract(db, contract_id=contract_id)

def delete_contract(db: Session, contract_id: int):
    db_contract = get_contract(db, contract_id=contract_id)
    if not db_contract: return None
    db.delete(db_contract)
    db.commit()

def create_contract_item(db: Session, item: schemas.ContractItemCreate, contract_id: int):
    db_item = models.ContractItem(**item.model_dump(), contract_id=contract_id)
    db.add(db_item); db.commit(); db.refresh(db_item); return db_item
def get_contract_item(db: Session, item_id: int):
    return db.query(models.ContractItem).options(joinedload(models.ContractItem.contract)).filter(models.ContractItem.id == item_id).first()
def update_contract_item(db: Session, item_id: int, item_update: schemas.ContractItemUpdate):
    db_item = get_contract_item(db, item_id=item_id)
    if not db_item: return None
    update_data = item_update.model_dump(exclude_none=True)
    for key, value in update_data.items(): setattr(db_item, key, value)
    db.add(db_item); db.commit(); db.refresh(db_item); return db_item
def delete_contract_item(db: Session, item_id: int):
    db_item = get_contract_item(db, item_id=item_id)
    if not db_item: return None
    db.delete(db_item); db.commit(); return db_item

def export_contracts_to_excel(db: Session, project_id: int):
    db_contracts = get_project_contracts(db, project_id=project_id)
    if not db_contracts:
        return io.BytesIO()

def import_contracts_from_excel(db: Session, project_id: int, file_contents: bytes):
    df = pd.read_excel(io.BytesIO(file_contents)).fillna('')
    created_count = 0
    updated_count = 0
    errors = []
    # Obtener el company_id desde el proyecto
    project = db.query(models.Project).filter(models.Project.id == project_id).first()
    if not project:
        return {"message": "Error fatal: Proyecto no encontrado.", "errors": ["El proyecto especificado no existe."]}
    company_id = project.company_id

    for index, row in df.iterrows():
        try:
            numero_contrato = str(row.get('Numero Contrato', '')).strip()
            razon_social_contratista = str(row.get('Contratista', '')).strip()
            if not numero_contrato or not razon_social_contratista:
                errors.append(f"Fila {index + 2}: Faltan 'Numero Contrato' o 'Contratista'.")
                continue
            contractor = get_contractor_by_razon_social(db, razon_social=razon_social_contratista, project_id=project_id)
            if not contractor:
                errors.append(f"Fila {index + 2}: No se encontró el contratista '{razon_social_contratista}'. Asegúrate de que exista en el sistema.")
                continue

            work_item_id = None
            work_item_description = str(row.get('Partida', '')).strip()
            if work_item_description:
                work_item = db.query(models.WorkItem).filter(models.WorkItem.project_id == project_id, models.WorkItem.description == work_item_description).first()
                if work_item:
                    work_item_id = work_item.id
                else:
                    errors.append(f"Fila {index + 2}: No se encontró la partida con descripción '{work_item_description}' en este proyecto.")

            # --- LÓGICA DE ACTUALIZAR O CREAR ---
            existing_contract = get_contract_by_number(db, project_id=project_id, numero_contrato=numero_contrato)

            if existing_contract:
                # Si existe, se actualiza
                update_data = schemas.ContractUpdate(
                    contractor_id=contractor.id,
                    trabajos=str(row.get('Trabajo', '')),
                    aplica_iva=str(row.get('Aplica IVA', 'SI')).strip().upper() == 'SI',
                    monto_contratado_manual=float(row.get('Monto Contratado (Manual)') or 0.0),
                    anticipo=float(row.get('Anticipo') or 0.0),
                    status=str(row.get('Status') or 'Borrador'),
                    external_url=str(row.get('URL Externa', '')),
                    work_item_id=work_item_id
                )
                update_contract(db, contract_id=existing_contract.id, contract_update=update_data)
                updated_count += 1
            else:
                # Si no existe, se crea
                contract_data = schemas.ContractCreate(
                    numero_contrato=numero_contrato, 
                    contractor_id=contractor.id, 
                    trabajos=str(row.get('Trabajo', '')), 
                    aplica_iva=str(row.get('Aplica IVA', 'SI')).strip().upper() == 'SI', 
                    monto_contratado_manual=float(row.get('Monto Contratado (Manual)') or 0.0), 
                    anticipo=float(row.get('Anticipo') or 0.0), 
                    status=str(row.get('Status') or 'Borrador'), 
                    external_url=str(row.get('URL Externa', '')), 
                    work_item_id=work_item_id,
                    start_date=row.get('Fecha Inicio') if not pd.isna(row.get('Fecha Inicio')) else None,
                    end_date=row.get('Fecha Fin') if not pd.isna(row.get('Fecha Fin')) else None,
                    avance_fisico=float(row.get('Avance Fisico') or 0.0)
                )
                create_contract(db=db, contract=contract_data, project_id=project_id)
                created_count += 1
        except Exception as e:
            errors.append(f"Fila {index + 2}: Error inesperado - {e}")
    return {"message": f"{created_count} contratos creados, {updated_count} actualizados.", "errors": errors}

    data_list = []
    for contract in db_contracts:
        schema_contract = schemas.Contract.from_orm(contract)
        data_list.append({
            "Numero Contrato": schema_contract.numero_contrato,
            "Contratista": schema_contract.contractor.razon_social,
            "Trabajos": schema_contract.trabajos,
            "Status": schema_contract.status,
            "Monto Contratado (Manual)": schema_contract.monto_contratado_manual,
            "Total Items (Calculado)": schema_contract.total_contratado_vigente,
            "IVA": schema_contract.iva,
            "Total con IVA": schema_contract.total_con_iva,
            "Anticipo": schema_contract.anticipo,
            "Total Estimado Acumulado": schema_contract.total_estimado_acumulado,
            "Avance Financiero (%)": schema_contract.progress,
        })

    df = pd.DataFrame(data_list)
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Contratos', index=False)
    output.seek(0)
    return output

def export_contract_items_to_excel(db: Session, contract_id: int):
    db_items = db.query(models.ContractItem).filter(models.ContractItem.contract_id == contract_id).all()
    if not db_items: return io.BytesIO()
    data_list = []
    for item in db_items:
        data_list.append({
            "Clave": item.clave, "Concepto": item.concepto, "Unidad": item.unidad,
            "Nivel/Zona": item.nivel_zona, "Tipo Concepto": item.tipo_concepto,
            "Avance Físico %": item.avance_fisico, "Precio Unitario": item.precio_unitario,
            "Cant. Contratada": item.cantidad_contratada, "Cant. Aditiva": item.cantidad_aditiva,
            "Cant. Deductiva": item.cantidad_deductiva, "parent_id": item.parent_id, "is_group": item.is_group
        })
    df = pd.DataFrame(data_list); output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer: df.to_excel(writer, sheet_name='CatalogoConceptos', index=False)
    output.seek(0); return output
def import_contract_items_from_excel(db: Session, contract_id: int, file_contents: bytes):
    workbook = openpyxl.load_workbook(io.BytesIO(file_contents))
    sheet = workbook.active
    
    created_count = 0
    errors = []
    
    clave_id_map = {
        item.clave: item.id 
        for item in db.query(models.ContractItem).filter(
            models.ContractItem.contract_id == contract_id, 
            models.ContractItem.clave != None
        ).all()
    }

    for i, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        try:
            # --- LECTURA DE COLUMNAS (Corregida a tu Excel) ---
            clave = str(row[0]) if row[0] else None
            concepto = str(row[1]) if row[1] else None
            unidad = str(row[2]) if row[2] else None
            nivel_zona = str(row[3]) if row[3] else None
            tipo_concepto_raw = str(row[4]).strip().title() if row[4] else "Ordinario"
            
            # --- ¡ORDEN CORREGIDO! ---
            cantidad_contratada = float(row[5]) if row[5] else 0.0
            precio_unitario = float(row[6]) if row[6] else 0.0
            # --- FIN DE LA CORRECCIÓN ---
            
            es_agrupador_raw = str(row[7]).strip().upper() if row[7] else "NO"
            parent_clave = str(row[8]) if row[8] else None
            # --- FIN DE LECTURA ---

            if not concepto:
                continue

            if tipo_concepto_raw not in ["Ordinario", "Extraordinario"]:
                tipo_concepto_raw = "Ordinario"
            
            is_group = True if es_agrupador_raw == "SI" else False
            
            # Busca el ID del padre en el mapa
            parent_id = clave_id_map.get(parent_clave) if parent_clave else None

            item_data = schemas.ContractItemCreate(
                clave=clave,
                concepto=concepto,
                unidad=unidad,
                nivel_zona=nivel_zona,
                tipo_concepto=tipo_concepto_raw,
                precio_unitario=precio_unitario,
                cantidad_contratada=cantidad_contratada,
                cantidad_aditiva=0.0,
                cantidad_deductiva=0.0,
                avance_fisico=0.0,
                parent_id=parent_id,
                is_group=is_group
            )
            
            db_item = create_contract_item(db=db, item=item_data, contract_id=contract_id)
            created_count += 1
            if clave:
                clave_id_map[clave] = db_item.id
            
        except Exception as e:
            errors.append(f"Fila {i}: Error inesperado - {e} (Fila: {row})")
            continue

    return {
        "message": f"{created_count} conceptos importados.",
        "errors": errors
    }

def create_estimate(db: Session, estimate: schemas.EstimateCreate, project_id: int):
    estimate_data = estimate.model_dump(exclude={'porcentaje_fondo_garantia'})
    db_estimate = models.Estimate(**estimate_data, project_id=project_id)
    if estimate.porcentaje_fondo_garantia is not None:
        monto_base = estimate.monto_estimado_manual or 0.0
        porcentaje = estimate.porcentaje_fondo_garantia / 100.0
        db_estimate.fondo_garantia = monto_base * porcentaje

    db.add(db_estimate)
    db.commit()
    db.refresh(db_estimate)
    return get_estimate(db, db_estimate.id)

def get_project_estimates(db: Session, project_id: int):
    return db.query(models.Estimate).filter(models.Estimate.project_id == project_id).options(
        joinedload(models.Estimate.contract), 
        joinedload(models.Estimate.estimate_items).options( 
            joinedload(models.EstimateItem.contract_item) 
        )
    ).all()
def get_estimate(db: Session, estimate_id: int):
    return db.query(models.Estimate).filter(models.Estimate.id == estimate_id).options(
        joinedload(models.Estimate.contract).options(
            joinedload(models.Contract.contractor),
            joinedload(models.Contract.contract_items)
        ),
        joinedload(models.Estimate.estimate_items).options(
            joinedload(models.EstimateItem.contract_item)
        )
    ).first()
def update_estimate(db: Session, estimate_id: int, estimate_update: schemas.EstimateUpdate):
    db_estimate = db.query(models.Estimate).filter(models.Estimate.id == estimate_id).first()
    if not db_estimate: return None

    porcentaje_fg = estimate_update.porcentaje_fondo_garantia
    update_data = estimate_update.model_dump(exclude_none=True, exclude={'porcentaje_fondo_garantia'})

    for key, value in update_data.items(): setattr(db_estimate, key, value)

    if porcentaje_fg is not None:
        total_items = db.query(func.sum(models.EstimateItem.cantidad_estimada * models.ContractItem.precio_unitario))\
            .join(models.ContractItem, models.EstimateItem.contract_item_id == models.ContractItem.id)\
            .filter(models.EstimateItem.estimate_id == estimate_id).scalar() or 0.0

        monto_base = total_items if total_items > 0 else (db_estimate.monto_estimado_manual or 0.0)
        porcentaje = porcentaje_fg / 100.0        
        db_estimate.fondo_garantia = monto_base * porcentaje
        
    db.add(db_estimate)
    db.commit()
    db.refresh(db_estimate)
    return get_estimate(db, estimate_id=estimate_id)

def partial_update_estimate(db: Session, db_estimate: models.Estimate, estimate_update: schemas.EstimatePartialUpdate):
    update_data = estimate_update.model_dump(exclude_none=True)
    for key, value in update_data.items():
        setattr(db_estimate, key, value)
    db.add(db_estimate); db.commit(); db.refresh(db_estimate)
    return db_estimate

def delete_estimate(db: Session, estimate_id: int):
    db_estimate = get_estimate(db, estimate_id=estimate_id)
    if not db_estimate: return None
    db.delete(db_estimate); db.commit(); return db_estimate
    
def export_estimates_to_excel(db: Session, project_id: int):
    db_estimates = get_project_estimates(db, project_id=project_id)
    if not db_estimates:
        return io.BytesIO()

    data_list = []
    for estimate in db_estimates:
        schema_estimate = schemas.Estimate.from_orm(estimate)
        data_list.append({
            "Numero Estimacion": schema_estimate.numero_estimacion,
            "Fecha": schema_estimate.fecha,
            "Numero Contrato": schema_estimate.contract.numero_contrato,
            "Subtotal": schema_estimate.subtotal,
            "Amortizacion Anticipo": schema_estimate.amortizacion_anticipo,
            "Fondo Garantia": schema_estimate.fondo_garantia,
            "Otras Retenciones": schema_estimate.otras_retenciones,
            "Otras Deductivas": schema_estimate.otras_deductivas,
            "Total a Pagar": schema_estimate.total_a_pagar,
        })

    df = pd.DataFrame(data_list)
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Estimaciones', index=False)
    output.seek(0)
    return output

def import_estimates_from_excel(db: Session, project_id: int, file_contents: bytes):
    df = pd.read_excel(io.BytesIO(file_contents)).fillna(0)
    created_count = 0
    errors = []

    for index, row in df.iterrows():
        try:
            contract = db.query(models.Contract).filter(models.Contract.numero_contrato == row['Numero Contrato'], models.Contract.project_id == project_id).first()
            if not contract:
                errors.append(f"Fila {index + 2}: No se encontró el contrato '{row['Numero Contrato']}' en este proyecto.")
                continue
            
            estimate_data = schemas.EstimateCreate(**row.to_dict(), contract_id=contract.id)
            create_estimate(db=db, estimate=estimate_data, project_id=project_id)
            created_count += 1
        except Exception as e:
            errors.append(f"Fila {index + 2}: Error inesperado - {e}")

    return {"message": f"{created_count} estimaciones creadas.", "errors": errors}

def get_total_estimated_quantity(db: Session, contract_item_id: int) -> float:
    total = db.query(func.sum(models.EstimateItem.cantidad_estimada)).filter(
        models.EstimateItem.contract_item_id == contract_item_id
    ).scalar()
    return total or 0.0
def create_estimate_item(db: Session, item: schemas.EstimateItemCreate, estimate_id: int):
    contract_item = db.query(models.ContractItem).filter(
        models.ContractItem.id == item.contract_item_id
    ).first()
    if not contract_item:
        raise HTTPException(status_code=404, detail="El concepto del contrato no existe.")
    if contract_item.is_group:
        raise HTTPException(status_code=400, detail="Error: No se puede estimar un 'agrupador'.")
    total_ya_estimado = get_total_estimated_quantity(db, item.contract_item_id)
    cantidad_total_vigente = (
        contract_item.cantidad_contratada + 
        contract_item.cantidad_aditiva - 
        contract_item.cantidad_deductiva
    )
    nueva_cantidad_total_estimada = total_ya_estimado + item.cantidad_estimada
    if nueva_cantidad_total_estimada > (cantidad_total_vigente + 0.01):
        faltante = cantidad_total_vigente - total_ya_estimado
        raise HTTPException(
            status_code=400,
            detail=f"Cantidad excede lo contratado. Vigente: {cantidad_total_vigente}, Estimado: {nueva_cantidad_total_estimada}, Faltante: {faltante:.2f}"
        )
    db_item = models.EstimateItem(**item.model_dump(), estimate_id=estimate_id)
    db.add(db_item); db.commit(); db.refresh(db_item); return db_item
def get_estimate_item(db: Session, item_id: int):
    return db.query(models.EstimateItem).filter(models.EstimateItem.id == item_id).first()
def update_estimate_item(db: Session, item_id: int, item_update: schemas.EstimateItemUpdate):
    db_item = get_estimate_item(db, item_id=item_id)
    if not db_item: return None
    update_data = item_update.model_dump(exclude_unset=True)
    for key, value in update_data.items(): setattr(db_item, key, value)
    db.add(db_item); db.commit(); db.refresh(db_item); return db_item
def delete_estimate_item(db: Session, item_id: int):
    db_item = get_estimate_item(db, item_id=item_id)
    if not db_item: return None
    db.delete(db_item); db.commit(); return db_item

def create_folder(db: Session, folder: schemas.FolderCreate, project_id: int):
    db_folder = models.Folder(name=folder.name, parent_id=folder.parent_id, project_id=project_id)
    db.add(db_folder); db.commit(); db.refresh(db_folder); return db_folder
def get_all_project_folders(db: Session, project_id: int):
    return db.query(models.Folder).filter(models.Folder.project_id == project_id).all()
def get_folder_contents(db: Session, folder_id: int):
    return db.query(models.Folder).filter(models.Folder.id == folder_id).options(
        joinedload(models.Folder.subfolders),
        joinedload(models.Folder.documents).joinedload(models.Document.versions)
    ).first()
def rename_folder(db: Session, folder_id: int, new_name: str):
    db_folder = db.query(models.Folder).filter(models.Folder.id == folder_id).first()
    if not db_folder: return None
    db_folder.name = new_name
    db.add(db_folder); db.commit(); db.refresh(db_folder); return db_folder
def delete_folder(db: Session, folder_id: int):
    db_folder = db.query(models.Folder).filter(models.Folder.id == folder_id).first()
    if not db_folder: raise HTTPException(status_code=404, detail="Carpeta no encontrada")
    if db.query(models.Folder).filter(models.Folder.parent_id == folder_id).first():
        raise HTTPException(status_code=400, detail="No se puede eliminar: La carpeta contiene subcarpetas.")
    if db.query(models.Document).filter(models.Document.folder_id == folder_id).first():
        raise HTTPException(status_code=400, detail="No se puede eliminar: La carpeta contiene documentos.")
    db.delete(db_folder); db.commit(); return True
def create_document_concept(db: Session, document: schemas.DocumentCreate):
    db_document = models.Document(name=document.name, folder_id=document.folder_id)
    db.add(db_document); db.commit(); db.refresh(db_document); return db_document
def _extract_ifc_properties(element: Any) -> dict:
    props = {}; props['Name'] = element.Name; props['IfcType'] = element.is_a()
    try:
        for rel in element.IsDefinedBy:
            if rel.is_a('IfcRelDefinesByProperties'):
                prop_def = rel.RelatingPropertyDefinition
                if prop_def.is_a('IfcPropertySet'):
                    pset_name = prop_def.Name; pset_props = {}
                    for prop in prop_def.HasProperties:
                        if prop.is_a('IfcPropertySingleValue'):
                            pset_props[prop.Name] = prop.NominalValue.wrappedValue
                    props[pset_name] = pset_props
                elif prop_def.is_a('IfcElementQuantity'):
                    qto_name = prop_def.Name; qto_props = {}
                    for quantity in prop_def.Quantities:
                        if quantity.is_a('IfcQuantityLength'): qto_props[quantity.Name] = quantity.LengthValue
                        elif quantity.is_a('IfcQuantityArea'): qto_props[quantity.Name] = quantity.AreaValue
                        elif quantity.is_a('IfcQuantityVolume'): qto_props[quantity.Name] = quantity.VolumeValue
                        elif quantity.is_a('IfcQuantityCount'): qto_props[quantity.Name] = quantity.CountValue
                    props[qto_name] = qto_props
    except Exception as e:
        print(f"Error extrayendo propiedades para {element.GlobalId}: {e}")
    return props
def create_new_document_version(db: Session, file: UploadFile, document_id: int):
    db_document = db.query(models.Document).filter(models.Document.id == document_id).first()
    if not db_document: return None 
    latest_version = db.query(func.max(models.DocumentVersion.version_number)).filter(
        models.DocumentVersion.document_id == document_id
    ).scalar() or 0
    new_version_number = latest_version + 1
    unique_filename = f"{uuid.uuid4()}_{file.filename}"
    file_path = os.path.join(UPLOAD_DIRECTORY, unique_filename)
    os.makedirs(UPLOAD_DIRECTORY, exist_ok=True)
    temp_file_path = f"{file_path}.tmp"
    ifc_file = None
    try:
        with open(temp_file_path, "wb") as buffer:
            shutil.copyfileobj(file.file, buffer)
        file_size = os.path.getsize(temp_file_path)
        if ifcopenshell and file.filename.lower().endswith(".ifc"):
            try:
                ifc_file = ifcopenshell.open(temp_file_path)
            except Exception as e:
                os.remove(temp_file_path) 
                raise HTTPException(status_code=400, detail=f"Archivo IFC inválido o corrupto: {e}")
        os.rename(temp_file_path, file_path)
    finally:
        file.file.close()
    db_version = models.DocumentVersion(
        document_id=document_id, version_number=new_version_number,
        filename=file.filename, file_path=file_path,
        file_type=file.content_type, file_size=file_size
    )
    db.add(db_version); db.commit(); db.refresh(db_version)
    
    try:
        from .utils import convert_ifc_to_frag
        if file.filename.lower().endswith(".ifc"):
            print(f"Iniciando conversión a Fragments para: {file_path}")
            convert_ifc_to_frag(file_path, UPLOAD_DIRECTORY)
    except ImportError:
        print("Warning: app.utils.convert_ifc_to_frag not found.")
    except Exception as e:
        print(f"Error triggering conversion: {e}")

    if ifc_file:
        print(f"Archivo IFC detectado. Extrayendo elementos para la versión {db_version.id}...")
        elements_to_add = []
        for element in ifc_file.by_type('IfcProduct'):
            if not element.GlobalId: continue 
            properties_dict = _extract_ifc_properties(element)
            db_element = models.BimElement(
                guid=element.GlobalId, ifc_type=element.is_a(),
                name=element.Name, properties=properties_dict,
                document_version_id=db_version.id 
            )
            elements_to_add.append(db_element)
        if elements_to_add:
            db.bulk_save_objects(elements_to_add)
            db.commit()
            print(f"Se extrajeron {len(elements_to_add)} elementos.")
    return db_version
def get_document_version(db: Session, version_id: int):
    return db.query(models.DocumentVersion).filter(models.DocumentVersion.id == version_id).first()
def delete_document_version(db: Session, version_id: int):
    db_version = db.query(models.DocumentVersion).filter(models.DocumentVersion.id == version_id).first()
    if not db_version: return None
    file_path = db_version.file_path
    db.delete(db_version); db.commit()
    if file_path and os.path.exists(file_path): os.remove(file_path)
    return {"ok": True}
def get_document_concept_with_project(db: Session, document_id: int):
    return db.query(models.Document).options(joinedload(models.Document.folder).joinedload(models.Folder.project)).filter(models.Document.id == document_id).first()
def delete_document_and_versions(db: Session, document_id: int):
    db_document = db.query(models.Document).options(joinedload(models.Document.versions)).filter(models.Document.id == document_id).first()
    if not db_document: return None
    for version in db_document.versions:
        if version.file_path and os.path.exists(version.file_path):
            os.remove(version.file_path)
    db.delete(db_document); db.commit(); return {"ok": True}
def link_document_to_contract(db: Session, document_id: int, contract_id: int):
    statement = models.document_contract_link.insert().values(document_id=document_id, contract_id=contract_id)
    db.execute(statement); db.commit(); return True
def link_document_to_work_item(db: Session, document_id: int, work_item_id: int):
    statement = models.document_workitem_link.insert().values(document_id=document_id, work_item_id=work_item_id)
    db.execute(statement); db.commit(); return True

def get_bcf_topics_by_project(db: Session, project_id: int):
    return db.query(models.BCFTopic).filter(models.BCFTopic.project_id == project_id).all()

def get_bcf_topic(db: Session, topic_guid: str):
    return db.query(models.BCFTopic).options(
        selectinload(models.BCFTopic.viewpoints).selectinload(models.BCFViewpoint.components),
        selectinload(models.BCFTopic.comments)
    ).filter(models.BCFTopic.guid == topic_guid).first()

def create_bcf_topic(db: Session, topic_data: schemas.BCFTopicCreate, project_id: int, author_email: str):
    db_topic = models.BCFTopic(
        **topic_data.model_dump(exclude={'viewpoints', 'comments'}),
        project_id=project_id,
        creation_author=author_email,
        modified_author=author_email
    )
    db.add(db_topic)
    db.flush()

    for viewpoint_data in topic_data.viewpoints:
        db_viewpoint = models.BCFViewpoint(
            **viewpoint_data.model_dump(exclude={'components'}),
            topic_guid=db_topic.guid
        )
        db.add(db_viewpoint)
        db.flush() 
        for component_data in viewpoint_data.components:
            db_component = models.BCFComponent(**component_data.model_dump(), viewpoint_guid=db_viewpoint.guid)
            db.add(db_component)

    for comment_data in topic_data.comments:
        db_comment = models.BCFComment(**comment_data.model_dump(), topic_guid=db_topic.guid, author=author_email)
        db.add(db_comment)
    db.commit()
    db.refresh(db_topic)
    return db_topic
def update_bcf_topic(db: Session, topic_guid: str, topic_update: schemas.BCFTopicUpdate, author_email: str):
    db_topic = get_bcf_topic(db, topic_guid=topic_guid)
    if not db_topic: return None
    update_data = topic_update.model_dump(exclude_none=True)
    for key, value in update_data.items(): setattr(db_topic, key, value)
    db_topic.modified_author = author_email
    db.add(db_topic); db.commit(); db.refresh(db_topic); return db_topic
def export_project_tasks_to_excel(db: Session, project_id: int):
    tasks = get_project_tasks(db, project_id)
    if not tasks:
        return io.BytesIO()

    data_list = []
    for task in tasks:
        total_cost = sum(tc.amount for tc in task.concepts)
        
        data_list.append({
            "ID": task.id,
            "Nombre": task.name,
            "Descripcion": task.description,
            "Fecha Inicio": task.start_date,
            "Fecha Fin": task.end_date,
            "Peso": task.weight,
            "Costo Estimado": total_cost,
            "Estado": task.status,
            "Progreso": task.progress,
            "ID Padre": task.parent_id
        })

    df = pd.DataFrame(data_list)
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Tareas", index=False)
    output.seek(0)
    return output


def import_tasks_from_excel(db: Session, project_id: int, file_contents: bytes):
    try:
        df = pd.read_excel(io.BytesIO(file_contents)).fillna("")
        # Ensure ID column is treated as string
        if "ID" in df.columns:
            df["ID"] = df["ID"].astype(str)
    except Exception as e:
        return {"error": f"Error al leer el archivo Excel: {str(e)}"}

    created_count = 0
    errors = []    
    wbs_map = {}
    project = db.query(models.Project).filter(models.Project.id == project_id).first()
    creator_id = None
    if project and project.company:
        user = db.query(models.User).filter(models.User.company_id == project.company_id).first()
        if user: creator_id = user.id

    if "ID" in df.columns:
        try:
            df = df.sort_values(by="ID")
        except:
            pass

    for index, row in df.iterrows():
        try:
            name = str(row.get("Nombre", "")).strip()
            if not name: continue

            wbs_id = str(row.get("ID", "")).strip()            
            parent_id = None
            parent_wbs_from_id_padre = str(row.get("ID Padre", "")).strip()
            if parent_wbs_from_id_padre and parent_wbs_from_id_padre in wbs_map:
                parent_id = wbs_map[parent_wbs_from_id_padre]
            elif parent_id is None and wbs_id and "." in wbs_id:
                parent_wbs_from_id = wbs_id.rsplit(".", 1)[0] 
                if parent_wbs_from_id in wbs_map:
                    parent_id = wbs_map[parent_wbs_from_id]

            start_date = row.get("Fecha Inicio")
            end_date = row.get("Fecha Fin")
            weight = row.get("Peso", 1.0)
            estimated_cost = row.get("Costo Estimado", 0.0)
            
            if pd.isna(start_date) or start_date == "": start_date = date.today()
            else: start_date = pd.to_datetime(start_date).date()
            
            if pd.isna(end_date) or end_date == "": end_date = date.today()
            else: end_date = pd.to_datetime(end_date).date()

            new_task = models.Task(
                name=name,
                description=str(row.get("Descripcion", "")),
                start_date=start_date,
                end_date=end_date,
                weight=float(weight) if weight != "" else 1.0,
                estimated_cost=float(estimated_cost) if estimated_cost != "" else 0.0,
                project_id=project_id,
                creator_id=creator_id,
                status="Pendiente",
                progress=0,
                parent_id=parent_id # Link to parent
            )
            db.add(new_task)
            db.flush() # Flush to get the ID
            
            if wbs_id:
                wbs_map[wbs_id] = new_task.id
                
            created_count += 1
        except Exception as e:
            errors.append(f"Fila {index + 2}: {str(e)}")
            continue

    db.commit()
    return {"message": f"Se importaron {created_count} tareas.", "errors": errors}

# --- Contract Item CRUD (Agregado/Redefinido para asegurar existencia) ---

def create_contract_item(db: Session, item: schemas.ContractItemCreate, contract_id: int):
    item_data = item.model_dump()
    db_item = models.ContractItem(**item_data, contract_id=contract_id)
    db.add(db_item)
    db.commit()
    db.refresh(db_item)
    return db_item

def get_contract_item(db: Session, item_id: int):
    return db.query(models.ContractItem).filter(models.ContractItem.id == item_id).first()

def update_contract_item(db: Session, item_id: int, item_update: schemas.ContractItemUpdate):
    db_item = get_contract_item(db, item_id=item_id)
    if not db_item:
        return None
    
    update_data = item_update.model_dump(exclude_unset=True)
    for key, value in update_data.items():
        setattr(db_item, key, value)
    
    db.add(db_item)
    db.commit()
    db.refresh(db_item)
    return db_item

def delete_contract_item(db: Session, item_id: int):
    db_item = get_contract_item(db, item_id=item_id)
    if not db_item:
        return None
    db.delete(db_item)
    db.commit()
    return db_item

def assign_cost_to_task(db: Session, project_id: int, task_id: int, amount: float, description: str):
    # 1. Buscar contrato "Borrador"
    draft_contract = db.query(models.Contract).filter(
        models.Contract.project_id == project_id,
        models.Contract.status == "Borrador"
    ).first()

    if not draft_contract:
        contractor = db.query(models.Contractor).filter(models.Contractor.project_id == project_id).first()
        if not contractor:
            contractor = models.Contractor(razon_social="Contratista General (Auto)", project_id=project_id)
            db.add(contractor); db.commit(); db.refresh(contractor)        
        draft_contract = models.Contract(
            numero_contrato=f"DRAFT-{project_id}-{int(date.today().strftime('%Y%m%d'))}",
            contractor_id=contractor.id,
            project_id=project_id,
            trabajos="Costos Directos de Tareas",
            status="Borrador"
        )
        db.add(draft_contract); db.commit(); db.refresh(draft_contract)
    
    new_item = models.ContractItem(
        contract_id=draft_contract.id,
        concepto=description,
        precio_unitario=amount,
        cantidad_contratada=1.0,
        task_id=task_id,
        tipo_concepto="Ordinario"
    )
    db.add(new_item)
    
    # 4. Actualizar costo estimado de la tarea (cache)
    task = db.query(models.Task).filter(models.Task.id == task_id).first()
    if task:
        task.estimated_cost = (task.estimated_cost or 0.0) + amount
        db.add(task)
        
    db.commit()
    db.refresh(new_item)
    return new_item

def get_task_costs(db: Session, task_id: int):
    """
    Obtiene el historial de costos asignados a una tarea.
    Retorna objetos ContractItem que luego el endpoint transformará o
    usará directamente si el schema coincide.
    """
    return db.query(models.ContractItem).options(
        joinedload(models.ContractItem.contract)
    ).filter(models.ContractItem.task_id == task_id).all()


